import json
import os
import re
from typing import List

from openpyxl import load_workbook

PATH = os.path.dirname(__file__) + r"\schedules"

GRADUATE_SCHOOL = 1

JSON_OUT = {}


def get_files(path: str) -> List:
    files = []

    for filename in os.listdir(path):

        f = os.path.join(path, filename)

        if os.path.isfile(f):
            files.append(f)

    return files


def get_merged_cells(cur_sheet, start_with_b) -> List:
    merged_cells = []

    for merged_cell in cur_sheet.merged_cells.ranges:
        if merged_cell.min_col == start_with_b + 2 and merged_cell.max_col == start_with_b + 2:
            merged_cells.append(merged_cell)

    return merged_cells


def is_merged_cell(cur_sheet, col, row) -> bool:
    is_merged = False

    for merged_cell in cur_sheet.merged_cells.ranges:
        if merged_cell.min_col == col and merged_cell.max_col == col \
                and (merged_cell.min_row == row or merged_cell.max_row == row):
            is_merged = True

    return is_merged


def get_search(arr: List, subject: str):
    out = None
    for item in arr:
        out = re.search(item, subject)
        if out:
            break
    return out


def get_audience(subject: str) -> str | None:
    subj_list = [r"ЛК-\d\d\d", r"ПА-\d\d", r"А-\d\d\d", r"У-\d\d\d", r"БЦ-\d\d\d"]
    subj_list_typos = [r"У - \d\d\d", r"У- \d\d\d", r"ЛК- \d\d\d", r"ЛК -\d\d\d", r"ПА- \d\d", r"А- \d\d\d",
                       r"ЛК \d\d\d", r"П-\d"]
    audience = get_search(subj_list, subject)

    if audience is None and "Иностранный" not in subject and "Физическая культура" not in subject \
            and re.search(r"БЦ-\d\d\d", subject) is None and re.search(r"ЦУВП-\d\d\d", subject) is None \
            and re.search(r"Библиотека", subject) is None \
            and subject != "Безопасность жизнедеятельности (Л 2-13н) Ларионов М.В." and subject != " ":

        subj_to_fix = get_search(subj_list_typos, subject)

        if subj_to_fix is not None:
            if re.search(r"ЛК \d\d\d", subject):
                aud = f"ЛК-{subject[subj_to_fix.end() - 3:subj_to_fix.end()]}"
            elif re.search(r"П-\d", subject):
                aud = f"ПА-{subject[subj_to_fix.end() - 1:subj_to_fix.end()]}"
            else:
                aud = subject[subj_to_fix.start():subj_to_fix.end()].replace(' ', '')
            subject = subject[0:subj_to_fix.start()] + aud + subject[subj_to_fix.end(): len(subject)]

            audience = get_search(subj_list, subject)

    if audience is not None:
        return subject[audience.start():audience.end()]

    return None


def parser(wb, i, week_parity_flag=False):
    sheet = wb.worksheets[i]
    start_with_b = 0

    if sheet.title == "3 курс":
        return

    if sheet.title == "ИГУиП 1 курс":
        start_with_b = 1

    MAX_COLS = sheet.max_column

    JSON_OUT[sheet.title] = {}

    merged_cells = get_merged_cells(sheet, start_with_b)

    if merged_cells:
        for day_coord in merged_cells:
            _, bottom, _, top = day_coord.bounds
            day_of_week = sheet[bottom][1 + start_with_b].value
            JSON_OUT[sheet.title][day_of_week] = {}
            delay = top - bottom + 1
            for i in range(delay):
                if i % 2 == 0:
                    period = sheet[bottom + i][2 + start_with_b].value
                    JSON_OUT[sheet.title][day_of_week][period] = {}

                    merged_rows = []
                    for k in range(MAX_COLS - 4):
                        is_merged = is_merged_cell(sheet, k + start_with_b + 5, bottom + i)

                        if is_merged:
                            merged_rows.append(k + start_with_b + 5)

                    for j in range(2):
                        week_parity = sheet[bottom + i + j][3 + start_with_b].value

                        if week_parity_flag:
                            if j == 0:
                                week_parity = "чёт"
                            else:
                                week_parity = "нечёт"

                        if week_parity == "чет":
                            week_parity = "чёт"
                        elif week_parity == "ЧЁТ.":
                            week_parity = "чёт"
                        elif week_parity == "НЕЧЁТ.":
                            week_parity = "нечёт"
                        elif week_parity == "нечет.":
                            week_parity = "нечёт"

                        JSON_OUT[sheet.title][day_of_week][period][week_parity] = {}

                        for k in range(MAX_COLS - 4 - start_with_b):
                            if j == 1 and (k + start_with_b + 5 in merged_rows):
                                subject = sheet[bottom + i][k + start_with_b + 4].value
                            else:
                                subject = sheet[bottom + i + j][k + start_with_b + 4].value

                            if subject is None:
                                continue

                            subject = subject.replace("\n", " ")

                            audience = get_audience(subject)

                            JSON_OUT[sheet.title][day_of_week][period][week_parity][subject.lower()] = audience


# итерация по path до каждого файла
for excel_path in get_files(PATH):
    # открывается файл Excel только в том случае, если файл не запущен, иначе выдается исключение
    try:
        wb = load_workbook(filename=excel_path)
    except PermissionError as e:
        print("Необходимо закрыть все файлы Excel для корректной работы.")

    # кол-во листов заносится в переменную
    quantity_sheets = len(wb.sheetnames)

    # данное условие необходимо для корректной обработки файла с расписанием аспирантуры
    if quantity_sheets == GRADUATE_SCHOOL:
        ws = wb.active

        if ws['D6'].value == "НЕД.":
            parser(wb, 0)
        else:
            parser(wb, 0, True)
    else:
        # парсим каждый лист
        for i in range(quantity_sheets):
            parser(wb, i)

# преобразуем и записываем словарь в JSON-объект и помещаем его в файл data_file.json
with open("data_file.json", "w") as write_file:
    json.dump(JSON_OUT, write_file, ensure_ascii=False)
